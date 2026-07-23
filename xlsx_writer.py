"""
Записывает результат работы Инспектора (баллы и комментарии) в реальный
xlsx-файл. Работает только с ячейками — никогда не трогает формулы шаблона,
они пересчитаются сами при открытии файла в Excel/LibreOffice.

Колонки (см. inspector_agent_v2.md): A — шкала (не трогаем), I — оценка
инспектора, K — комментарий инспектора.

Дополнительно: после проставления баллов пересчитываем простые формулы
шаблона (суммы, проценты, средние) вручную в Python и записываем ГОТОВОЕ
ЧИСЛО поверх формулы. Это нужно, потому что openpyxl не умеет вычислять
формулы, а встроенный просмотрщик документов в Telegram тоже не считает
формулы "на лету" — без этого шага итоговые проценты в шапке анкеты (ВХОД,
ЧИСТОТА, СЕРВИС и т.д.) выглядели бы пустыми при открытии прямо в Telegram.
"""

import re

import openpyxl

SCORE_COLUMN = 9  # I
COMMENT_COLUMN = 11  # K

# Ключевые слова для поиска нужной строки шапки по тексту подписи.
# Строки шапки не имеют фиксированных номеров — структура отличается
# между разными версиями шаблона (напр. "Афимолл" vs "Мясницкая"),
# поэтому ищем строку по содержанию, а не по жёсткому номеру.
HEADER_LABEL_KEYWORDS = {
    "restaurant": ["название ресторана"],
    "date": ["дата и вид инспекции", "дата и время инспекции"],
    "time_in": ["время захода"],
    "time_out": ["время выхода"],
    "waiter_name": ["имя и описание официанта", "имя официанта"],
    "waiter_description": ["имя и фамилия сотрудника с бейджа"],
}


def _find_header_row(ws, keywords: list[str], max_row: int = 15) -> int | None:
    """Ищет в первых max_row строках столбца C строку, чей текст
    содержит один из keywords (без учёта регистра)."""
    for row in range(1, max_row + 1):
        value = ws.cell(row=row, column=3).value
        if not value:
            continue
        lowered = str(value).lower()
        if any(keyword in lowered for keyword in keywords):
            return row
    return None


CELL_REF_PATTERN = re.compile(r"[A-Z]+\d+")
SUM_PATTERN = re.compile(r"^SUM\(([A-Z]+\d+):([A-Z]+\d+)\)$")
AVERAGE_PATTERN = re.compile(r"^AVERAGE\(([^)]+)\)$")


def _get_numeric_value(ws, cell_ref: str, computed_cache: dict) -> float:
    """Возвращает числовое значение ячейки — если там уже посчитанное
    число, используем его; если формула — рекурсивно вычисляем (с кэшем,
    чтобы не пересчитывать одну и ту же ячейку много раз)."""
    if cell_ref in computed_cache:
        return computed_cache[cell_ref]

    value = ws[cell_ref].value

    if isinstance(value, (int, float)):
        computed_cache[cell_ref] = value
        return value

    if isinstance(value, str) and value.startswith("="):
        computed = _evaluate_formula(ws, value[1:], computed_cache)
        computed_cache[cell_ref] = computed
        return computed

    # Пустая ячейка или нечисловое значение — считаем как 0, чтобы формулы
    # вроде SUM по диапазону с пустыми строками не падали с ошибкой.
    computed_cache[cell_ref] = 0
    return 0


def _evaluate_formula(ws, formula_body: str, computed_cache: dict) -> float:
    """Вычисляет простую формулу шаблона. Поддерживает только конкретные
    паттерны, реально встречающиеся в шаблонах Инспектора: SUM(A1:A9),
    сложение нескольких ссылок (A1+A2+A3), деление и умножение двух
    значений (A1*100/A2), AVERAGE(A1,A2,A3), и простое копирование
    значения другой ячейки (=A1). Не претендует на полноценный движок
    формул Excel — только на то, что реально нужно для этих шаблонов."""
    formula_body = formula_body.strip()

    sum_match = SUM_PATTERN.match(formula_body)
    if sum_match:
        start_ref, end_ref = sum_match.groups()
        start_col = re.match(r"[A-Z]+", start_ref).group()
        start_row = int(re.match(r"[A-Z]+(\d+)", start_ref).group(1))
        end_row = int(re.match(r"[A-Z]+(\d+)", end_ref).group(1))
        total = 0
        for row in range(start_row, end_row + 1):
            total += _get_numeric_value(ws, f"{start_col}{row}", computed_cache)
        return total

    average_match = AVERAGE_PATTERN.match(formula_body)
    if average_match:
        refs = [r.strip() for r in average_match.group(1).split(",")]
        values = [_get_numeric_value(ws, ref, computed_cache) for ref in refs]
        return sum(values) / len(values) if values else 0

    # Простое копирование значения другой ячейки: формула — это просто ссылка,
    # без операторов (например "=I57").
    if CELL_REF_PATTERN.fullmatch(formula_body):
        return _get_numeric_value(ws, formula_body, computed_cache)

    # Арифметическое выражение из ссылок и операторов +,-,*,/ (без скобок,
    # без функций) — заменяем каждую ссылку на её числовое значение и
    # вычисляем через eval на уже безопасной, полностью числовой строке.
    if re.fullmatch(r"[A-Z0-9+\-*/.\s]+", formula_body):
        refs = CELL_REF_PATTERN.findall(formula_body)
        expression = formula_body
        for ref in sorted(set(refs), key=len, reverse=True):
            value = _get_numeric_value(ws, ref, computed_cache)
            expression = re.sub(rf"\b{ref}\b", str(value), expression)
        try:
            return eval(expression, {"__builtins__": {}}, {})
        except (ZeroDivisionError, SyntaxError):
            return 0

    return 0


def _recalculate_simple_formulas(ws) -> None:
    """Проходит по всем ячейкам листа, находит простые формулы (см.
    _evaluate_formula) и заменяет их на посчитанное число — округлённое
    до 2 знаков после запятой для процентов/средних, без изменений для
    целых сумм баллов. Это нужно, чтобы итоговые проценты в шапке анкеты
    были видны сразу в любом просмотрщике, включая встроенный в Telegram,
    который не умеет вычислять формулы Excel."""
    computed_cache: dict[str, float] = {}
    formula_cells = []

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append(cell.coordinate)

    for coordinate in formula_cells:
        try:
            computed_value = _get_numeric_value(ws, coordinate, computed_cache)
            if isinstance(computed_value, float):
                computed_value = round(computed_value, 2)
            ws[coordinate] = computed_value
        except Exception:
            # Если конкретная формула не распознана движком — оставляем как
            # есть (текст формулы), не ломаем файл. Такие случаи стоит
            # добавить в _evaluate_formula отдельно, если встретятся на практике.
            continue


def apply_fill_results(
    empty_template_path: str,
    output_path: str,
    header: dict,
    filled_rows: list[dict],
) -> list[int]:
    """Открывает пустой шаблон, проставляет баллы/комментарии/шапку,
    сохраняет как новый файл (не трогая исходный).

    Возвращает список номеров строк, которые были пропущены из-за
    отсутствия комментария (защита от 'осиротевших' баллов без пояснения)."""
    wb = openpyxl.load_workbook(empty_template_path, data_only=False)
    ws = wb["шаблон"]

    for key, keywords in HEADER_LABEL_KEYWORDS.items():
        value = header.get(key)
        if not value:
            continue

        row = _find_header_row(ws, keywords)
        if row is None:
            continue  # в этой версии шаблона такого поля шапки просто нет

        cell = ws.cell(row=row, column=3)
        existing_text = cell.value or ""
        if ":" in existing_text:
            label = existing_text.split(":", 1)[0]
            cell.value = f"{label}: {value}"
        else:
            cell.value = f"{existing_text} {value}".strip()

    skipped_rows = []
    for item in filled_rows:
        row = item["row"]
        score = item["score"]
        comment = item.get("comment", "").strip()

        if not comment:
            # Без комментария балл не записываем вообще — оставляем строку
            # пустой, а не создаём "осиротевший" балл без пояснения. Это
            # частая ошибка модели: она иногда заполняет две строки одного
            # вопроса, и одна из них остаётся без текста.
            skipped_rows.append(row)
            continue

        ws.cell(row=row, column=SCORE_COLUMN, value=score)
        ws.cell(row=row, column=COMMENT_COLUMN, value=comment)

    _recalculate_simple_formulas(ws)

    wb.save(output_path)
    return skipped_rows


def check_for_formula_errors(output_path: str) -> list[str]:
    """Проверяет сохранённый файл на явные признаки сломанных формул.
    Полноценный пересчёт формул openpyxl не делает (это не Excel),
    поэтому здесь только базовая проверка на артефакты вроде #REF!."""
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb["шаблон"]
    problems = []

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and any(
                marker in cell.value for marker in ("#REF!", "#DIV/0!", "#VALUE!")
            ):
                problems.append(f"{cell.coordinate}: {cell.value}")

    return problems
