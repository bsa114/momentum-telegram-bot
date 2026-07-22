"""
Извлекает из пустой xlsx-анкеты (шаблон Steak it Easy) структуру,
понятную модели: разделы, пункты, блоки шкалы с их строками.
Это заменяет ручное чтение файла моделью — мы даём ей уже готовый,
компактный JSON вместо необходимости "смотреть" в реальный xlsx.
"""

import openpyxl


def extract_template_structure(xlsx_path: str, sheet_name: str = "шаблон") -> dict:
    """Возвращает структуру анкеты: шапку и список оцениваемых блоков.

    Формат блока:
    {
        "block_start_row": int,      # первая строка блока (название пункта, если есть отдельно)
        "section": str,              # текущий раздел (последний встреченный заголовок раздела)
        "rows": [
            {"row": int, "scale_value": float, "description": str}
        ]
    }
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name]

    # Шапка — это всё, что идёт до первой строки с реальным баллом в столбце A.
    # Ищем первую такую строку, чтобы не зависеть от фиксированных номеров,
    # которые отличаются между версиями шаблона.
    first_scored_row = None
    for row in range(1, ws.max_row + 1):
        a_value = ws.cell(row=row, column=1).value
        if isinstance(a_value, (int, float)):
            first_scored_row = row
            break
    if first_scored_row is None:
        first_scored_row = ws.max_row + 1  # на всякий случай, если баллов вообще нет

    header_info = {}
    for row in range(1, first_scored_row):
        c_value = ws.cell(row=row, column=3).value
        if c_value:
            header_info[f"row_{row}"] = str(c_value)

    blocks = []
    current_block = None
    current_section = None

    for row in range(first_scored_row, ws.max_row + 1):
        a_value = ws.cell(row=row, column=1).value
        c_value = ws.cell(row=row, column=3).value

        if a_value is None and c_value is not None and not current_block:
            # похоже на заголовок раздела или пункта (текст без баллов рядом)
            text = str(c_value).strip()
            if text.isupper() or len(text) < 60:
                current_section = text

        if a_value is not None and c_value is not None:
            # Строки с формулами (например "=A37+A39+...") — это итог по разделу,
            # не обычный пункт анкеты для заполнения. Их пропускаем.
            if isinstance(a_value, str) and a_value.strip().startswith("="):
                current_block = None
                continue

            if current_block is None:
                current_block = {
                    "section": current_section,
                    "rows": [],
                }
                blocks.append(current_block)
            try:
                scale_value = float(a_value)
            except (TypeError, ValueError):
                scale_value = a_value
            current_block["rows"].append(
                {"row": row, "scale_value": scale_value, "description": str(c_value).strip()}
            )
        elif current_block is not None and a_value is None and c_value is None:
            # пустая строка — блок закончился
            current_block = None

    return {"header": header_info, "blocks": blocks}


def build_compact_template_text(structure: dict) -> str:
    """Собирает компактное текстовое представление структуры для промпта —
    легче для модели читать, чем вложенный JSON, и экономит токены."""
    lines = ["ШАПКА АНКЕТЫ:"]
    for key, value in structure["header"].items():
        lines.append(f"  {key}: {value}")

    lines.append("\nПУНКТЫ АНКЕТЫ (раздел / строка / балл / описание варианта):")
    current_section = None
    for block in structure["blocks"]:
        section = block["section"] or "(без раздела)"
        if section != current_section:
            lines.append(f"\n--- Раздел: {section} ---")
            current_section = section
        for r in block["rows"]:
            lines.append(f"  строка {r['row']}: балл {r['scale_value']} — {r['description']}")

    return "\n".join(lines)
